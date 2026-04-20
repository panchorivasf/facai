import sys, csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

tmp_dir   = sys.argv[1]
out_path  = sys.argv[2]

def read_csv(name):
    with open(f"{tmp_dir}/{name}") as f:
        return list(csv.DictReader(f))

meta   = {r["key"]: r["value"] for r in read_csv("meta.csv")}
status = read_csv("status.csv")
flags  = read_csv("flags.csv")
dbh    = read_csv("dbh.csv")
growth = read_csv("growth.csv")

wb = Workbook()

DARK  = "2E4057"
GREY  = "F5F5F5"
WHITE = "FFFFFF"
WARN  = "FFF9C4"
CRIT  = "FFDDC1"
INFO  = "E8F5E9"

thin   = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def hdr_font():  return Font(name="Arial", bold=True, color="FFFFFF", size=10)
def body_font(): return Font(name="Arial", size=10)
def hdr_fill():  return PatternFill("solid", fgColor=DARK)
def alt_fill(i): return PatternFill("solid", fgColor=(GREY if i % 2 == 0 else WHITE))

def write_sheet(wb, title, headers, rows, col_widths=None, row_fills=None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = hdr_font()
        cell.fill      = hdr_fill()
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = border
    for i, row in enumerate(rows, start=2):
        ws.append(row)
        fill = (row_fills[i - 2] if row_fills else alt_fill(i))
        for cell in ws[i]:
            cell.font      = body_font()
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(vertical="center")
    if col_widths:
        for idx, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = w
    ws.freeze_panes = "A2"
    return ws

# ── Overview sheet ────────────────────────────────────────────────────────────
ws_ov = wb.active
ws_ov.title = "Overview"
ov_data = [
    ["Dataset",    meta["dataset_name"]],
    ["Rows",       meta["n_rows"]],
    ["Trees",      meta["n_trees"]],
    ["Plots",      meta["n_plots"]],
    ["Year min",   meta["yr_min"]],
    ["Year max",   meta["yr_max"]],
    ["Verdict",    meta["verdict"]],
]
for i, (k, v) in enumerate(ov_data, start=1):
    kc = ws_ov.cell(row=i, column=1, value=k)
    vc = ws_ov.cell(row=i, column=2, value=v)
    kc.font   = Font(name="Arial", bold=True, size=10)
    vc.font   = body_font()
    kc.border = border
    vc.border = border
    kc.fill   = alt_fill(i)
    vc.fill   = alt_fill(i)
ws_ov.column_dimensions["A"].width = 18
ws_ov.column_dimensions["B"].width = 40

# ── Status sheet ──────────────────────────────────────────────────────────────
write_sheet(wb, "Status",
            ["Status", "Label", "Count", "Pct (%)"],
            [[r["Status"], r["Label"], r["n"], r["pct"]] for r in status],
            col_widths=[10, 28, 12, 10])

# ── DBH sheet ─────────────────────────────────────────────────────────────────
d_keys = list(dbh[0].keys())
write_sheet(wb, "DBH",
            d_keys,
            [[r[k] for k in d_keys] for r in dbh],
            col_widths=[12] * len(d_keys))

# ── Growth sheet ──────────────────────────────────────────────────────────────
g_keys = list(growth[0].keys())
write_sheet(wb, "Growth",
            g_keys,
            [[r[k] for k in g_keys] for r in growth],
            col_widths=[12] * len(g_keys))

# ── Flags sheet ───────────────────────────────────────────────────────────────
sev_fills = {
    "critical": PatternFill("solid", fgColor=CRIT),
    "warning":  PatternFill("solid", fgColor=WARN),
    "info":     PatternFill("solid", fgColor=INFO),
}
row_fills = [sev_fills.get(r["Severity"], alt_fill(i))
             for i, r in enumerate(flags)]
write_sheet(wb, "Flags",
            ["Flag", "Count", "Severity"],
            [[r["Flag"], r["Count"], r["Severity"]] for r in flags],
            col_widths=[36, 10, 12],
            row_fills=row_fills)

wb.save(out_path)
